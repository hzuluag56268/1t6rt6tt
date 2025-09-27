import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Tuple, Set
import os


class AsignadorTurnosSencillos:
    """
    Asigna turnos "MANR", "TANR", "MASR", "TASR", "ASIG" con estas reglas:
    - Decisión por día según personal disponible/turnos operativos:
      * 9-10 personal → NO asignar turnos en ese día
      * 11 personal → asignar "MANR" a un trabajador y "TANR" a otro trabajador
      * 12 personal → asignar "MANR", "TANR", "MASR", "TASR" (no se asigna "ASIG")
      * 13+ personal → asignar "MANR", "TANR", "MASR", "TASR", "ASIG"
    - Solo se asigna entre trabajadores elegibles: ['PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
      'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE']
    - No asignar si ese día ya existe un turno "BLPTD" o "NANRD"
    - Equidad: los trabajadores deben tener la misma cantidad de turnos de cada tipo en lo posible
    - Colores: rojo medio oscuro para "MASR", "TASR". rojo claro para "MANR", "TANR". Gris medio para "ASIG"
    - Archivo de entrada: "horarioUnificado_con_mofis.xlsx"
    - Archivo de salida: "horarioUnificado_con_sencillos.xlsx"
    - Conserva la hoja de Estadísticas sin modificar para compatibilidad con stat_transformada.py
    """

    TRABAJADORES_ELEGIBLES = [
        'PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
        'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE'
    ]

    # Trabajadores para días con conflictos BLPTD/NANRD
    TRABAJADORES_PREFERIDOS_CONFLICTOS = ["YIS", "MAQ", "DJO", "AFG", "JLF", "JMV"]
    TRABAJADORES_ALTERNATIVOS_CONFLICTOS = ['FCE', 'JBV', 'GCE', 'GMT', 'HZG', 'JIS', 'CDT', 'WGG']
    TRABAJADORES_SEGUNDO_GRUPO_CONFLICTOS = ['HLG', 'ECE', 'DFB', 'MLS', 'FCE', 'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE']

    # Colores para los turnos
    COLOR_MANR = "FF9999"    # Rojo claro
    COLOR_TANR = "FF9999"    # Rojo claro  
    COLOR_MASR = "CC0000"    # Rojo medio oscuro
    COLOR_TASR = "CC0000"    # Rojo medio oscuro
    COLOR_ASIG = "808080"    # Gris medio
    
    # Colores para turnos de conflictos
    COLOR_MLPR = "FFD700"    # Dorado
    COLOR_TLPR = "FFD700"    # Dorado
    COLOR_TLPT = "FFD700"    # Dorado
    COLOR_TANT = "FFA500"    # Naranja
    COLOR_MAST = "FF4500"    # Naranja rojizo

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_mofis.xlsx",
            "horarioUnificado_con_diurnas.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            elegido = "horarioUnificado_con_mofis.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()

        # Snapshot del estado original
        self.original_nonempty: Set[Tuple[int, int]] = set()
        self._snapshot_estado_original()

        # Contadores de equidad por tipo de turno
        self.contador_manr: Dict[str, int] = defaultdict(int)
        self.contador_tanr: Dict[str, int] = defaultdict(int)
        self.contador_masr: Dict[str, int] = defaultdict(int)
        self.contador_tasr: Dict[str, int] = defaultdict(int)
        self.contador_asig: Dict[str, int] = defaultdict(int)
        
        # Contadores para turnos de conflictos
        self.contador_mlpr: Dict[str, int] = defaultdict(int)
        self.contador_tlpr: Dict[str, int] = defaultdict(int)
        self.contador_tlpt: Dict[str, int] = defaultdict(int)
        self.contador_tant: Dict[str, int] = defaultdict(int)
        self.contador_mast: Dict[str, int] = defaultdict(int)

        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _snapshot_estado_original(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is not None and str(valor).strip() != "":
                    self.original_nonempty.add((fila, col))

    def _es_celda_originalmente_vacia(self, fila: int, col: int) -> bool:
        return (fila, col) not in self.original_nonempty

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _contar_personal_operativo(self, col_dia: int) -> int:
        """Cuenta el personal operativo usando la misma lógica que procesador_horarios.py"""
        turnos_no_operativos_y_mofis = {
            # Turnos básicos
            "DESC", "TROP",
            # Turnos completos
            "VACA", "COME", "COMT", "COMS",
            # Turnos adicionales originales
            "SIND", "CMED", "CERT", "LICR",
            # Formación, instrucción y entrenamiento
            "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
            "MENT", "TENT", "NENT", "AENT",
            "MINS", "TINS", "NINS", "AINS",
            # Gestión, oficinas y grupos de trabajo
            "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
            "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
            "MGST", "TGST", "MOFI", "TOFI",
            # Operativos y asignaciones especiales
            "CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X",
            # Turnos MOFIS adicionales
            "MN", "TN", "MS", "TS", "N", "S"
        }
        
        count = 0
        for fila in range(2, 26):  # Filas 2-25
            cell_value = self.ws.cell(row=fila, column=col_dia).value
            if cell_value is None or str(cell_value).strip() == "":
                count += 1  # Celda vacía = operativo
            else:
                valor_limpio = str(cell_value).strip().upper()
                if valor_limpio not in turnos_no_operativos_y_mofis:
                    count += 1  # No está en lista de no operativos = operativo

        return count

    def _existe_turno_conflictivo_en_dia(self, col_dia: int) -> bool:
        """Verifica si ya existe BLPTD o NANRD en el día"""
        for fila in range(2, 26):
            v = self.ws.cell(row=fila, column=col_dia).value
            if v is None:
                continue
            val = str(v).strip().upper()
            if val in {"BLPTD", "NANRD"}:
                return True
        return False

    def _existe_turno_repetido_en_dia(self, turno: str, col_dia: int) -> bool:
        """Verifica si ya existe el turno específico en el día"""
        for fila in range(2, 26):
            v = self.ws.cell(row=fila, column=col_dia).value
            if v is None:
                continue
            val = str(v).strip().upper()
            if val == turno.upper():
                return True
        return False

    def _obtener_trabajadores_disponibles(self, col_dia: int) -> List[str]:
        disponibles: List[str] = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if (
                (celda.value is None or str(celda.value).strip() == "")
                and self._es_celda_originalmente_vacia(fila, col_dia)
            ):
                disponibles.append(trabajador)
        return disponibles

    def _inicializar_contadores_desde_hoja(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            trabajador = str(trabajador).strip().upper()
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val == "MANR":
                    self.contador_manr[trabajador] += 1
                elif val == "TANR":
                    self.contador_tanr[trabajador] += 1
                elif val == "MASR":
                    self.contador_masr[trabajador] += 1
                elif val == "TASR":
                    self.contador_tasr[trabajador] += 1
                elif val == "ASIG":
                    self.contador_asig[trabajador] += 1
                elif val == "MLPR":
                    self.contador_mlpr[trabajador] += 1
                elif val == "TLPR":
                    self.contador_tlpr[trabajador] += 1
                elif val == "TLPT":
                    self.contador_tlpt[trabajador] += 1
                elif val == "TANT":
                    self.contador_tant[trabajador] += 1
                elif val == "MAST":
                    self.contador_mast[trabajador] += 1

    def _seleccionar_equitativo_por_tipo(self, candidatos: List[str], tipo_turno: str) -> Optional[str]:
        if not candidatos:
            return None
        
        contador = None
        if tipo_turno == "MANR":
            contador = self.contador_manr
        elif tipo_turno == "TANR":
            contador = self.contador_tanr
        elif tipo_turno == "MASR":
            contador = self.contador_masr
        elif tipo_turno == "TASR":
            contador = self.contador_tasr
        elif tipo_turno == "ASIG":
            contador = self.contador_asig
        elif tipo_turno == "MLPR":
            contador = self.contador_mlpr
        elif tipo_turno == "TLPR":
            contador = self.contador_tlpr
        elif tipo_turno == "TLPT":
            contador = self.contador_tlpt
        elif tipo_turno == "TANT":
            contador = self.contador_tant
        elif tipo_turno == "MAST":
            contador = self.contador_mast
        else:
            return random.choice(candidatos)
        
        # Equidad por menor conteo del tipo específico
        min_val = min(contador[c] for c in candidatos)
        empatados = [c for c in candidatos if contador[c] == min_val]
        return random.choice(empatados)

    def _actualizar_contadores(self, trabajador: str, tipo_turno: str, delta: int = 1) -> None:
        if tipo_turno == "MANR":
            self.contador_manr[trabajador] += delta
        elif tipo_turno == "TANR":
            self.contador_tanr[trabajador] += delta
        elif tipo_turno == "MASR":
            self.contador_masr[trabajador] += delta
        elif tipo_turno == "TASR":
            self.contador_tasr[trabajador] += delta
        elif tipo_turno == "ASIG":
            self.contador_asig[trabajador] += delta
        elif tipo_turno == "MLPR":
            self.contador_mlpr[trabajador] += delta
        elif tipo_turno == "TLPR":
            self.contador_tlpr[trabajador] += delta
        elif tipo_turno == "TLPT":
            self.contador_tlpt[trabajador] += delta
        elif tipo_turno == "TANT":
            self.contador_tant[trabajador] += delta
        elif tipo_turno == "MAST":
            self.contador_mast[trabajador] += delta

    def _asignar_turno(self, trabajador: str, col_dia: int, tipo_turno: str) -> bool:
        """Asigna un turno específico a un trabajador"""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if not self._es_celda_originalmente_vacia(fila, col_dia):
            return False
        
        self.ws.cell(row=fila, column=col_dia, value=tipo_turno)
        celda = self.ws.cell(row=fila, column=col_dia)
        
        # Aplicar colores según el tipo de turno
        if tipo_turno == "MANR":
            celda.fill = PatternFill(start_color=self.COLOR_MANR, end_color=self.COLOR_MANR, fill_type="solid")
        elif tipo_turno == "TANR":
            celda.fill = PatternFill(start_color=self.COLOR_TANR, end_color=self.COLOR_TANR, fill_type="solid")
        elif tipo_turno == "MASR":
            celda.fill = PatternFill(start_color=self.COLOR_MASR, end_color=self.COLOR_MASR, fill_type="solid")
        elif tipo_turno == "TASR":
            celda.fill = PatternFill(start_color=self.COLOR_TASR, end_color=self.COLOR_TASR, fill_type="solid")
        elif tipo_turno == "ASIG":
            celda.fill = PatternFill(start_color=self.COLOR_ASIG, end_color=self.COLOR_ASIG, fill_type="solid")
        elif tipo_turno == "MLPR":
            celda.fill = PatternFill(start_color=self.COLOR_MLPR, end_color=self.COLOR_MLPR, fill_type="solid")
        elif tipo_turno == "TLPR":
            celda.fill = PatternFill(start_color=self.COLOR_TLPR, end_color=self.COLOR_TLPR, fill_type="solid")
        elif tipo_turno == "TLPT":
            celda.fill = PatternFill(start_color=self.COLOR_TLPT, end_color=self.COLOR_TLPT, fill_type="solid")
        elif tipo_turno == "TANT":
            celda.fill = PatternFill(start_color=self.COLOR_TANT, end_color=self.COLOR_TANT, fill_type="solid")
        elif tipo_turno == "MAST":
            celda.fill = PatternFill(start_color=self.COLOR_MAST, end_color=self.COLOR_MAST, fill_type="solid")
        
        self._actualizar_contadores(trabajador, tipo_turno, 1)
        return True

    def _puede_asignar_turnos(self, col_dia: int) -> Tuple[bool, int, int, str]:
        """
        Verifica si se pueden asignar turnos en un día.
        Retorna (puede_asignar, personal_operativo, disponibles_count, razon)
        """
        if self._existe_turno_conflictivo_en_dia(col_dia):
            return False, 0, 0, "Turno conflictivo (BLPTD/NANRD)"
            
        personal_operativo = self._contar_personal_operativo(col_dia)
        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        disponibles_count = len(disponibles)
        
        if personal_operativo <= 10:
            return False, personal_operativo, disponibles_count, "Poco personal (≤10)"
        elif personal_operativo == 11 and disponibles_count < 2:
            return False, personal_operativo, disponibles_count, "Pocos disponibles para 11 personal"
        elif personal_operativo == 12 and disponibles_count < 4:
            return False, personal_operativo, disponibles_count, "Pocos disponibles para 12 personal"
        elif personal_operativo >= 13 and disponibles_count < 5:
            return False, personal_operativo, disponibles_count, "Pocos disponibles para 13+ personal"
        else:
            return True, personal_operativo, disponibles_count, "OK"

    def asignar_turnos_en_dia(self, col_dia: int) -> List[str]:
        """
        Asigna turnos en un día según las reglas.
        Retorna lista de asignaciones realizadas
        """
        puede_asignar, personal_operativo, _, _ = self._puede_asignar_turnos(col_dia)
        if not puede_asignar:
            return []

        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        asignaciones = []

        if personal_operativo == 11:
            # Asignar MANR y TANR
            turnos_a_asignar = ["MANR", "TANR"]
        elif personal_operativo == 12:
            # Asignar MANR, TANR, MASR, TASR (no ASIG)
            turnos_a_asignar = ["MANR", "TANR", "MASR", "TASR"]
        elif personal_operativo >= 13:
            # Asignar MANR, TANR, MASR, TASR, ASIG
            turnos_a_asignar = ["MANR", "TANR", "MASR", "TASR", "ASIG"]
        else:
            return []

        # Verificar que no existan turnos repetidos y asignar
        for turno in turnos_a_asignar:
            if self._existe_turno_repetido_en_dia(turno, col_dia):
                continue  # Saltar si ya existe este turno
                
            # Seleccionar trabajador equitativo para este tipo de turno
            elegido = self._seleccionar_equitativo_por_tipo(disponibles, turno)
            if elegido and self._asignar_turno(elegido, col_dia, turno):
                asignaciones.append(f"{turno}→{elegido}")
                disponibles.remove(elegido)

        return asignaciones

    def _obtener_trabajadores_disponibles_conflictos(self, col_dia: int, lista_trabajadores: List[str]) -> List[str]:
        """Obtiene trabajadores disponibles para días con conflictos"""
        disponibles: List[str] = []
        for trabajador in lista_trabajadores:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if (
                (celda.value is None or str(celda.value).strip() == "")
                and self._es_celda_originalmente_vacia(fila, col_dia)
            ):
                disponibles.append(trabajador)
        return disponibles

    def asignar_turnos_conflictos_primer_grupo(self, col_dia: int) -> List[str]:
        """Asigna turnos MLPR, TLPR, TLPT en días con conflictos BLPTD/NANRD"""
        asignaciones = []
        turnos_primer_grupo = ["MLPR", "TLPR", "TLPT"]
        
        # Obtener trabajadores disponibles preferidos y alternativos
        disponibles_preferidos = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_PREFERIDOS_CONFLICTOS)
        disponibles_alternativos = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_ALTERNATIVOS_CONFLICTOS)
        
        # Determinar qué lista usar: solo alternativos si NO hay preferidos disponibles
        if disponibles_preferidos:
            trabajadores_a_usar = disponibles_preferidos
        else:
            trabajadores_a_usar = disponibles_alternativos
        
        # Asignar cada turno si hay trabajadores disponibles y el turno no existe
        for turno in turnos_primer_grupo:
            if not self._existe_turno_repetido_en_dia(turno, col_dia) and trabajadores_a_usar:
                elegido = self._seleccionar_equitativo_por_tipo(trabajadores_a_usar, turno)
                if elegido and self._asignar_turno(elegido, col_dia, turno):
                    asignaciones.append(f"{turno}→{elegido}")
                    trabajadores_a_usar.remove(elegido)
        
        return asignaciones

    def asignar_turnos_conflictos_segundo_grupo(self, col_dia: int) -> List[str]:
        """Asigna turnos MANR, TANR, TANT, MAST, MASR, TASR en días con conflictos"""
        asignaciones = []
        turnos_segundo_grupo = ["MANR", "TANR", "TANT", "MAST", "MASR", "TASR"]
        
        # Obtener trabajadores disponibles del segundo grupo
        disponibles = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_SEGUNDO_GRUPO_CONFLICTOS)
        
        # Asignar cada turno si hay trabajadores disponibles y el turno no existe
        for turno in turnos_segundo_grupo:
            if not self._existe_turno_repetido_en_dia(turno, col_dia) and disponibles:
                elegido = self._seleccionar_equitativo_por_tipo(disponibles, turno)
                if elegido and self._asignar_turno(elegido, col_dia, turno):
                    asignaciones.append(f"{turno}→{elegido}")
                    disponibles.remove(elegido)
        
        return asignaciones

    def asignar_turnos_en_dia_con_conflictos(self, col_dia: int) -> List[str]:
        """Asigna turnos en días con conflictos BLPTD/NANRD"""
        asignaciones_totales = []
        
        # Diagnóstico: obtener información del día
        header = self.ws.cell(row=1, column=col_dia).value
        
        # Primer grupo: MLPR, TLPR, TLPT
        disponibles_preferidos = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_PREFERIDOS_CONFLICTOS)
        disponibles_alternativos = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_ALTERNATIVOS_CONFLICTOS)
        disponibles_segundo_grupo = self._obtener_trabajadores_disponibles_conflictos(col_dia, self.TRABAJADORES_SEGUNDO_GRUPO_CONFLICTOS)
        
        print(f"\n🔍 DIAGNÓSTICO {header}:")
        print(f"   Preferidos disponibles: {len(disponibles_preferidos)} → {disponibles_preferidos}")
        print(f"   Alternativos disponibles: {len(disponibles_alternativos)} → {disponibles_alternativos}")
        print(f"   Segundo grupo disponibles: {len(disponibles_segundo_grupo)} → {disponibles_segundo_grupo}")
        
        asignaciones_primer_grupo = self.asignar_turnos_conflictos_primer_grupo(col_dia)
        asignaciones_totales.extend(asignaciones_primer_grupo)
        
        # Segundo grupo: MANR, TANR, TANT, MAST, MASR, TASR
        asignaciones_segundo_grupo = self.asignar_turnos_conflictos_segundo_grupo(col_dia)
        asignaciones_totales.extend(asignaciones_segundo_grupo)
        
        print(f"   Primer grupo asignado: {len(asignaciones_primer_grupo)}/3 → {asignaciones_primer_grupo}")
        print(f"   Segundo grupo asignado: {len(asignaciones_segundo_grupo)}/6 → {asignaciones_segundo_grupo}")
        print(f"   TOTAL ASIGNADO: {len(asignaciones_totales)}/9")
        
        return asignaciones_totales

    def _actualizar_fila_conteo_operativo(self) -> None:
        """Actualiza la fila de conteo operativo estático usando la misma lógica que procesador_horarios.py"""
        turnos_no_operativos_y_mofis = {
            # Turnos básicos
            "DESC", "TROP",
            # Turnos completos
            "VACA", "COME", "COMT", "COMS",
            # Turnos adicionales originales
            "SIND", "CMED", "CERT", "LICR",
            # Formación, instrucción y entrenamiento
            "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
            "MENT", "TENT", "NENT", "AENT",
            "MINS", "TINS", "NINS", "AINS",
            # Gestión, oficinas y grupos de trabajo
            "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
            "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
            "MGST", "TGST", "MOFI", "TOFI",
            # Operativos y asignaciones especiales
            "CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X",
            # Turnos MOFIS adicionales
            "MN", "TN", "MS", "TS", "N", "S"
        }
        
        # Buscar la fila de conteo operativo estático
        fila_conteo = None
        for fila in range(1, self.ws.max_row + 1):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip() == "TURNOS OPERATIVOS":
                fila_conteo = fila
                break
        
        if fila_conteo is None:
            print("⚠️  No se encontró la fila 'TURNOS OPERATIVOS' para actualizar")
            return
        
        # Actualizar conteos para cada columna
        for col in range(2, self.ws.max_column + 1):
            conteo_operativos = 0
            # Contar turnos operativos según la lógica correcta
            for row in range(2, 26):  # Filas 2-25
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value is None or str(cell_value).strip() == "":
                    conteo_operativos += 1
                else:
                    valor_limpio = str(cell_value).strip().upper()
                    if valor_limpio not in turnos_no_operativos_y_mofis:
                        conteo_operativos += 1
            
            # Escribir el conteo actualizado
            celda_conteo = self.ws.cell(row=fila_conteo, column=col)
            celda_conteo.value = conteo_operativos
            
            # Aplicar colores según el conteo (misma lógica que procesador_horarios.py)
            rojo_intenso = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            rojo_medio = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
            azul_clarito = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
            verde_clarito = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            verde_intenso = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
            sin_relleno = PatternFill(fill_type=None)
            fuente_blanca = Font(color="FFFFFF")
            
            if conteo_operativos <= 8:
                celda_conteo.fill = rojo_intenso
                celda_conteo.font = fuente_blanca
            elif conteo_operativos == 9:
                celda_conteo.fill = rojo_medio
            elif conteo_operativos == 10:
                celda_conteo.fill = azul_clarito
            elif conteo_operativos == 11:
                celda_conteo.fill = verde_clarito
            elif conteo_operativos == 12:
                celda_conteo.fill = verde_intenso
            else:
                celda_conteo.fill = sin_relleno
        
        print("✅ Fila de conteo operativo estático actualizada")

    def _copiar_hoja_estadisticas(self) -> None:
        """Copia la hoja de Estadísticas sin modificar para mantener compatibilidad"""
        if "Estadísticas" in self.wb.sheetnames:
            print("✅ Conservando hoja de Estadísticas original para compatibilidad con stat_transformada.py")
        else:
            print("⚠️  No se encontró la hoja de Estadísticas en el archivo de entrada")

    def _generar_reporte_detallado(self) -> None:
        """Genera un reporte detallado de disponibilidad y asignaciones por día"""
        print("\n" + "="*80)
        print("REPORTE DETALLADO DE ASIGNACIÓN DE TURNOS SENCILLOS")
        print("="*80)
        
        max_col = self.ws.max_column
        asignaciones_realizadas = []
        asignaciones_conflictos = []
        dias_sin_asignar = []
        dias_con_conflictos = []
        
        for col in range(2, max_col + 1):
            header = self.ws.cell(row=1, column=col).value
            if not header or header == "SIGLA ATCO":
                continue
                
            # Verificar si se puede asignar
            puede_asignar, personal_operativo, disponibles_count, razon = self._puede_asignar_turnos(col)
            
            if not puede_asignar:
                if razon.startswith("Turno conflictivo"):
                    # Intentar asignar turnos para días con conflictos
                    asignaciones_conflicto = self.asignar_turnos_en_dia_con_conflictos(col)
                    if asignaciones_conflicto:
                        asignaciones_conflictos.append((header, asignaciones_conflicto, personal_operativo, disponibles_count))
                    else:
                        dias_con_conflictos.append((header, personal_operativo, disponibles_count, razon))
                else:
                    dias_sin_asignar.append((header, personal_operativo, disponibles_count, razon))
            else:
                # Intentar asignar turnos normales
                asignaciones = self.asignar_turnos_en_dia(col)
                if asignaciones:
                    asignaciones_realizadas.append((header, asignaciones, personal_operativo, disponibles_count))
        
        # Mostrar asignaciones realizadas
        print(f"\n🎯 ASIGNACIONES REALIZADAS: {len(asignaciones_realizadas)}")
        if asignaciones_realizadas:
            print("-" * 80)
            for header, asignaciones, personal, disponibles in asignaciones_realizadas:
                print(f"{header:>8}: {', '.join(asignaciones):40} (Personal: {personal:2d}, Disponibles: {disponibles})")
        
        # Mostrar asignaciones en días con conflictos
        print(f"\n🔥 ASIGNACIONES EN DÍAS CON CONFLICTOS: {len(asignaciones_conflictos)}")
        if asignaciones_conflictos:
            print("-" * 80)
            for header, asignaciones, personal, disponibles in asignaciones_conflictos:
                print(f"{header:>8}: {', '.join(asignaciones):60} (Conflicto BLPTD/NANRD)")
        
        # Mostrar días sin asignar
        if dias_sin_asignar:
            print(f"\n⚠️  DÍAS SIN ASIGNAR: {len(dias_sin_asignar)}")
            print("-" * 80)
            for header, personal, disponibles, razon in dias_sin_asignar:
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles:2d} - {razon}")
        
        # Mostrar días con conflictos
        if dias_con_conflictos:
            print(f"\n🚫 DÍAS CON CONFLICTOS: {len(dias_con_conflictos)}")
            print("-" * 80)
            for header, personal, disponibles, razon in dias_con_conflictos:
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles:2d} - {razon}")
        
        # Resumen de equidad
        print(f"\n📊 RESUMEN DE EQUIDAD POR TRABAJADOR:")
        print("-" * 100)
        print(f"{'Trabajador':>10} {'MANR':>5} {'TANR':>5} {'MASR':>5} {'TASR':>5} {'ASIG':>5} {'MLPR':>5} {'TLPR':>5} {'TLPT':>5} {'TANT':>5} {'MAST':>5} {'TOTAL':>6}")
        print("-" * 100)
        
        # Mostrar todos los trabajadores que pueden tener turnos
        todos_trabajadores = set(self.TRABAJADORES_ELEGIBLES + self.TRABAJADORES_PREFERIDOS_CONFLICTOS + self.TRABAJADORES_ALTERNATIVOS_CONFLICTOS + self.TRABAJADORES_SEGUNDO_GRUPO_CONFLICTOS)
        
        for trabajador in sorted(todos_trabajadores):
            if self._obtener_fila_trabajador(trabajador):
                manr = self.contador_manr[trabajador]
                tanr = self.contador_tanr[trabajador]
                masr = self.contador_masr[trabajador]
                tasr = self.contador_tasr[trabajador]
                asig = self.contador_asig[trabajador]
                mlpr = self.contador_mlpr[trabajador]
                tlpr = self.contador_tlpr[trabajador]
                tlpt = self.contador_tlpt[trabajador]
                tant = self.contador_tant[trabajador]
                mast = self.contador_mast[trabajador]
                total = manr + tanr + masr + tasr + asig + mlpr + tlpr + tlpt + tant + mast
                if total > 0:  # Solo mostrar trabajadores con turnos asignados
                    print(f"{trabajador:>10} {manr:>5} {tanr:>5} {masr:>5} {tasr:>5} {asig:>5} {mlpr:>5} {tlpr:>5} {tlpt:>5} {tant:>5} {mast:>5} {total:>6}")
        
        print("="*100)
        return len(asignaciones_realizadas) + len(asignaciones_conflictos)

    def procesar_todos_los_dias(self) -> None:
        print(f"📁 Procesando archivo: {self.archivo_entrada}")
        
        # Actualizar la fila de conteo operativo estático antes de asignar
        print("🔄 Actualizando fila de conteo operativo estático...")
        self._actualizar_fila_conteo_operativo()
        
        # Generar reporte detallado
        num_asignaciones = self._generar_reporte_detallado()
        
        # Conservar hoja de estadísticas
        self._copiar_hoja_estadisticas()

        # Guardar archivo
        salida = "horarioUnificado_con_sencillos.xlsx"
        try:
            self.wb.save(salida)
            print(f"\n✅ Archivo guardado como: {salida}")
            print(f"📊 Total de asignaciones realizadas: {num_asignaciones}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"\n✅ Archivo por defecto en uso. Guardado como: {alternativo}")
            print(f"📊 Total de asignaciones realizadas: {num_asignaciones}")


if __name__ == "__main__":
    asignador = AsignadorTurnosSencillos()
    asignador.procesar_todos_los_dias()

