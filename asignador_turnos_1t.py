import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Tuple
import os
from openpyxl.comments import Comment


class AsignadorTurnos:
    """
    Asigna turnos "1T" (1 hora extra) o "7" (1 hora extra + 6 horas adicionales)
    sobre el archivo procesado, con las siguientes reglas:
    - Decisión por día según personal disponible (fila cuyo encabezado en la columna A es "TURNOS OPERATIVOS"):
      * ≤8  → no asignar
      * =9  → asignar turno "7"
      * ≥10 → asignar turno "1T"
    - Solo se asigna entre trabajadores elegibles: ['GCE', 'YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']
    - Prioridad para quienes tuvieron DESC/TROP/SIND el día anterior
    - Evitar, en lo posible, asignar si el día anterior tuvo "1T" o "7"
    - Restricción dura: NO asignar si el día anterior tuvo BANTD, BLPTD, NLPRD, NANRD, 1T o 7
    - Restricción blanda: Evitar si el día anterior tuvo NANTD o NLPTD
    - Verificar que en ese día NO exista ya un turno "1T" ni "7" ni "BLPTD" ni "BANTD" (no duplicar)
    - No asignar si el día siguiente tiene BANTD, BLPTD, 1T o 7
    - Restricción Torre: para "GCE", solo asignar 1T si el conteo de la fila "Torre" es ≤3
    - Equidad:
      * Grupo 1T: cuenta "1T" + "7" (toda persona con 1 hora extra)
      * Grupo 6RT: cuenta solo "7" (6 horas adicionales)
      * Para turno "1T": balancear por el grupo 1T
      * Para turno "7": balancear por el grupo 1T; si hay empate, usar grupo 6RT como desempate
    - Actualiza hoja "Estadísticas" con columnas: SIGLA, DESC, 1T (1T+7), 6RT (solo 7)
    """

    TRABAJADORES_ELEGIBLES = ['GCE', 'YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']

    def __init__(self, archivo_procesado: Optional[str] = None) -> None:
        self.archivo_procesado = self._resolver_archivo_entrada(archivo_procesado)
        self.wb = openpyxl.load_workbook(self.archivo_procesado)
        self.ws = self._obtener_hoja_horario()
        self.contador_grupo_1t: Dict[str, int] = defaultdict(int)
        self.contador_grupo_6rt: Dict[str, int] = defaultdict(int)
        # Inicializar contadores a partir de asignaciones ya existentes
        self._inicializar_contadores_desde_hoja()
        # Lista para almacenar información de días no asignados
        self.dias_no_asignados: List[Dict] = []
        # Crear el patrón de relleno naranja claro para turnos 1T/7
        self.fill_naranja_claro = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        random.seed()

    def _resolver_archivo_entrada(self, preferido: Optional[str]) -> str:
        candidatos = [
            preferido,
            "horario_procesado_con_sabados_domingos.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        for c in [c for c in candidatos if c]:
            if os.path.exists(c):
                return c
        return "horario_procesado_con_sabados_domingos.xlsx"

    def _obtener_hoja_horario(self):
        """
        Obtiene la hoja principal de horario (no la de estadísticas).
        Si existe una hoja llamada "Estadísticas", se omite.
        """
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        # Por seguridad, si solo existe "Estadísticas", usarla igualmente
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        """Devuelve la fila (int) donde está el trabajador en la columna A, o None si no se encuentra."""
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _obtener_nombre_dia(self, col_dia: int) -> str:
        """Obtiene el nombre del día desde el encabezado de la columna."""
        header = self.ws.cell(row=1, column=col_dia).value
        if header:
            return str(header)
        return f"Columna {col_dia}"

    def _obtener_valor_dia_anterior(self, trabajador: str, col_dia: int) -> str:
        """Obtiene el valor del día anterior para un trabajador."""
        if col_dia <= 2:
            return "N/A"
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return "N/A"
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        return str(valor) if valor is not None else "Vacío"

    def _obtener_valor_dia_siguiente(self, trabajador: str, col_dia: int) -> str:
        """Obtiene el valor del día siguiente para un trabajador."""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return "N/A"
        if col_dia + 1 > self.ws.max_column:
            return "N/A"
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        return str(valor) if valor is not None else "Vacío"

    def _tiene_prioridad_dia_anterior(self, trabajador: str, col_dia: int) -> bool:
        """True si el día anterior el trabajador tiene DESC, TROP o SIND."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"DESC", "TROP", "SIND"}

    def _tuvo_extra_dia_anterior(self, trabajador: str, col_dia: int) -> bool:
        """True si el día anterior el trabajador tuvo 1T o 7."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"1T", "7"}

    def _tuvo_restriccion_dura_ayer(self, trabajador: str, col_dia: int) -> bool:
        """True si ayer tuvo BANTD, BLPTD, NLPRD, NANRD, 1T o 7."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"BANTD", "BLPTD", "NLPRD", "NANRD", "1T", "7"}

    def _tiene_restriccion_dura_manana(self, trabajador: str, col_dia: int) -> bool:
        """True si mañana tiene BANTD, BLPTD, 1T o 7."""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"BANTD", "BLPTD", "1T", "7"}

    def _tuvo_restriccion_blanda_ayer(self, trabajador: str, col_dia: int) -> bool:
        """True si ayer tuvo NANTD o NLPTD (evitar si es posible)."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"NANTD", "NLPTD"}

    def _obtener_trabajadores_disponibles(self, col_dia: int) -> List[str]:
        """Devuelve los trabajadores elegibles cuya celda del día está vacía."""
        disponibles = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if celda.value is None or str(celda.value).strip() == "":
                disponibles.append(trabajador)
        return disponibles

    def _obtener_conteo_operativos(self, col_dia: int) -> Optional[int]:
        """Busca la fila con etiqueta 'TURNOS OPERATIVOS' en la columna A y devuelve el entero de esa columna del día."""
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TURNOS OPERATIVOS":
                valor = self.ws.cell(row=fila, column=col_dia).value
                try:
                    return int(valor)
                except Exception:
                    return None
        return None

    def _obtener_conteo_torre(self, col_dia: int) -> Optional[int]:
        """Busca la fila con etiqueta 'Torre' (columna A) y devuelve el entero de esa columna del día."""
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TORRE":
                valor = self.ws.cell(row=fila, column=col_dia).value
                try:
                    return int(valor)
                except Exception:
                    return None
        return None

    def _determinar_turno_por_personal(self, col_dia: int) -> Optional[str]:
        """
        - ≤8: None (no asignar)
        - =9: "7"
        - ≥10: "1T"
        """
        disponible = self._obtener_conteo_operativos(col_dia)
        if disponible is None:
            return None
        if disponible <= 8:
            return None
        if disponible == 9:
            return "7"
        return "1T"

    def _existe_turno_1t_o_7_en_dia(self, col_dia: int) -> bool:
        """True si en ese día ya existe un 1T, 7, BLPTD o BANTD en cualquier trabajador (filas 2-25)."""
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            val = str(valor).strip().upper()
            if val in {"1T", "7", "BLPTD", "BANTD"}:
                return True
        return False

    def _seleccionar_equitativo(self, candidatos: List[str], turno: str) -> Optional[str]:
        """
        Para turno "1T": usar contador_grupo_1t.
        Para turno "7": usar contador_grupo_1t y, como desempate, contador_grupo_6rt.
        """
        if not candidatos:
            return None

        # Contador principal siempre es el de 1T (porque 7 también suma al grupo 1T)
        principal = self.contador_grupo_1t
        min_principal = min(principal[c] for c in candidatos)
        candidatos_min = [c for c in candidatos if principal[c] == min_principal]

        if turno == "7" and len(candidatos_min) > 1:
            # Desempate con grupo 6RT
            secundaria = self.contador_grupo_6rt
            min_sec = min(secundaria[c] for c in candidatos_min)
            candidatos_min = [c for c in candidatos_min if secundaria[c] == min_sec]

        return random.choice(candidatos_min)

    def _actualizar_contadores(self, trabajador: str, turno: str) -> None:
        if turno == "1T":
            self.contador_grupo_1t[trabajador] += 1
        elif turno == "7":
            self.contador_grupo_1t[trabajador] += 1
            self.contador_grupo_6rt[trabajador] += 1

    def _inicializar_contadores_desde_hoja(self) -> None:
        """Inicializa los contadores 1T (1T+7) y 6RT (solo 7) leyendo asignaciones ya presentes."""
        max_col = self.ws.max_column
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val == "1T":
                    self.contador_grupo_1t[trabajador] += 1
                elif val == "7":
                    self.contador_grupo_1t[trabajador] += 1
                    self.contador_grupo_6rt[trabajador] += 1

    def _marcar_alerta_restriccion_dura(self, col_dia: int, mensaje: str = "Bloqueado por restricción dura (BANTD/BLPTD/NLPTD/NANRD)") -> None:
        """Agrega un comentario en el encabezado del día para alertar restricción dura."""
        header_cell = self.ws.cell(row=1, column=col_dia)
        texto_existente = header_cell.comment.text if header_cell.comment else ""
        nuevo_texto = (texto_existente + "\n" if texto_existente else "") + mensaje
        header_cell.comment = Comment(nuevo_texto, "Asignador")

    def _analizar_razones_no_asignacion(self, col_dia: int, turno_objetivo: str) -> Dict[str, List[str]]:
        """Analiza las razones por las que cada trabajador no puede ser asignado."""
        razones = {}
        nombre_dia = self._obtener_nombre_dia(col_dia)
        
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            razones_trabajador = []
            
            # Verificar si el trabajador existe
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                razones_trabajador.append("Trabajador no encontrado en la hoja")
                razones[trabajador] = razones_trabajador
                continue
            
            # Verificar si la celda está ocupada
            celda = self.ws.cell(row=fila, column=col_dia)
            if celda.value is not None and str(celda.value).strip() != "":
                razones_trabajador.append(f"Celda ocupada: '{celda.value}'")
                razones[trabajador] = razones_trabajador
                continue
            
            # Verificar restricción dura del día anterior
            if self._tuvo_restriccion_dura_ayer(trabajador, col_dia):
                valor_ayer = self._obtener_valor_dia_anterior(trabajador, col_dia)
                razones_trabajador.append(f"Restricción dura ayer: '{valor_ayer}' (BANTD/BLPTD/NLPRD/NANRD/1T/7)")
            
            # Verificar restricción dura del día siguiente
            if self._tiene_restriccion_dura_manana(trabajador, col_dia):
                valor_manana = self._obtener_valor_dia_siguiente(trabajador, col_dia)
                razones_trabajador.append(f"Restricción dura mañana: '{valor_manana}' (BANTD/BLPTD/1T/7)")
            
            # Verificar restricción Torre para GCE
            if trabajador == "GCE" and turno_objetivo == "1T":
                torre = self._obtener_conteo_torre(col_dia)
                if torre is not None and torre > 3:
                    razones_trabajador.append(f"Restricción Torre: conteo {torre} > 3")
            
            # Verificar restricción blanda del día anterior
            if self._tuvo_restriccion_blanda_ayer(trabajador, col_dia):
                valor_ayer = self._obtener_valor_dia_anterior(trabajador, col_dia)
                razones_trabajador.append(f"Restricción blanda ayer: '{valor_ayer}' (NANTD/NLPTD)")
            
            # Verificar si tuvo extra ayer
            if self._tuvo_extra_dia_anterior(trabajador, col_dia):
                valor_ayer = self._obtener_valor_dia_anterior(trabajador, col_dia)
                razones_trabajador.append(f"Tuvo extra ayer: '{valor_ayer}' (1T/7)")
            
            # Si no hay razones, el trabajador está disponible
            if not razones_trabajador:
                razones_trabajador.append("Disponible")
            
            razones[trabajador] = razones_trabajador
        
        return razones

    def _mostrar_alerta_dia_no_asignado(self, col_dia: int, turno_objetivo: str, razones: Dict[str, List[str]]) -> None:
        """Muestra una alerta detallada para un día no asignado."""
        nombre_dia = self._obtener_nombre_dia(col_dia)
        operativos = self._obtener_conteo_operativos(col_dia)
        
        print(f"\n{'='*80}")
        print(f"⚠️  DÍA NO ASIGNADO: {nombre_dia} (Columna {col_dia})")
        print(f"{'='*80}")
        print(f"Turno objetivo: {turno_objetivo}")
        print(f"Personal operativo: {operativos}")
        print(f"Regla aplicada: {'≤8: no asignar' if operativos <= 8 else '=9: asignar 7' if operativos == 9 else '≥10: asignar 1T'}")
        
        # Verificar si ya existe un turno 1T/7 en el día
        if self._existe_turno_1t_o_7_en_dia(col_dia):
            print(f"\n RAZÓN PRINCIPAL: Ya existe un turno 1T/7/BLPTD/BANTD en este día")
            return
        
        print(f"\n📋 ANÁLISIS POR TRABAJADOR:")
        print(f"{'-'*80}")
        
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            razones_trabajador = razones.get(trabajador, ["Error en análisis"])
            estado = "✅ DISPONIBLE" if razones_trabajador == ["Disponible"] else "❌ NO DISPONIBLE"
            
            print(f"{trabajador:4} | {estado}")
            for razon in razones_trabajador:
                if razon != "Disponible":
                    print(f"      └─ {razon}")
        
        # Almacenar información para resumen final
        self.dias_no_asignados.append({
            'dia': nombre_dia,
            'columna': col_dia,
            'turno_objetivo': turno_objetivo,
            'operativos': operativos,
            'ya_existe_turno': self._existe_turno_1t_o_7_en_dia(col_dia),
            'razones': razones
        })

    def asignar_turno_en_dia(self, col_dia: int) -> Optional[str]:
        """Intenta asignar "1T" o "7" en el día (columna) indicado, retornando el trabajador o None."""
        turno = self._determinar_turno_por_personal(col_dia)
        if turno is None:
            # Analizar por qué no se asigna (personal insuficiente)
            operativos = self._obtener_conteo_operativos(col_dia)
            nombre_dia = self._obtener_nombre_dia(col_dia)
            print(f"\n{'='*60}")
            print(f"ℹ️  DÍA SIN ASIGNACIÓN: {nombre_dia} (Columna {col_dia})")
            print(f"{'='*60}")
            print(f"Personal operativo: {operativos}")
            print(f"Razón: Personal insuficiente (≤8 operativos)")
            return None

        # Verificar que ese día NO haya ya un 1T ni un 7 ni BLPTD ni BANTD
        if self._existe_turno_1t_o_7_en_dia(col_dia):
            # Analizar razones detalladas aunque ya exista turno
            razones = self._analizar_razones_no_asignacion(col_dia, turno)
            self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
            return None

        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        if not disponibles:
            # Analizar razones detalladas
            razones = self._analizar_razones_no_asignacion(col_dia, turno)
            self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
            return None

        tenia_disponibles = bool(disponibles)

        # Excluir por restricción dura del día anterior
        disponibles_antes = disponibles.copy()
        disponibles = [t for t in disponibles if not self._tuvo_restriccion_dura_ayer(t, col_dia)]
        if not disponibles:
            if tenia_disponibles:
                self._marcar_alerta_restriccion_dura(col_dia, "Bloqueado por restricción dura (ayer: BANTD/BLPTD/NLPRD/NANRD/1T/7)")
                # Analizar razones detalladas
                razones = self._analizar_razones_no_asignacion(col_dia, turno)
                self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
            return None

        # Excluir por restricción dura del día siguiente
        disponibles = [t for t in disponibles if not self._tiene_restriccion_dura_manana(t, col_dia)]
        if not disponibles:
            if tenia_disponibles:
                self._marcar_alerta_restriccion_dura(col_dia, "Bloqueado por restricción dura (mañana: BANTD/BLPTD/1T/7)")
                # Analizar razones detalladas
                razones = self._analizar_razones_no_asignacion(col_dia, turno)
                self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
            return None

        # Restricción Torre para GCE cuando turno objetivo es 1T
        if turno == "1T":
            torre = self._obtener_conteo_torre(col_dia)
            if torre is not None and torre > 3 and "GCE" in disponibles:
                disponibles = [t for t in disponibles if t != "GCE"]
                if not disponibles:
                    # Analizar razones detalladas
                    razones = self._analizar_razones_no_asignacion(col_dia, turno)
                    self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
                    return None

        # Prioridades con restricciones (blandas y prioridad DESC/TROP/SIND de ayer)
        nivel1 = []
        nivel2 = []
        nivel3 = []
        nivel4 = []
        nivel5 = []
        for t in disponibles:
            prioridad = self._tiene_prioridad_dia_anterior(t, col_dia)
            tuvo_extra = self._tuvo_extra_dia_anterior(t, col_dia)
            blanda = self._tuvo_restriccion_blanda_ayer(t, col_dia)
            if prioridad and not tuvo_extra and not blanda:
                nivel1.append(t)
            elif not tuvo_extra and not blanda:
                nivel2.append(t)
            elif prioridad and blanda:
                nivel3.append(t)
            elif blanda:
                nivel4.append(t)
            else:
                nivel5.append(t)

        for candidatos in (nivel1, nivel2, nivel3, nivel4, nivel5):
            elegido = self._seleccionar_equitativo(candidatos, turno)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                # Asignar el turno y aplicar formato naranja claro
                self.ws.cell(row=fila, column=col_dia, value=turno)
                self._aplicar_formato_turno(fila, col_dia, turno)
                self._actualizar_contadores(elegido, turno)
                nombre_dia = self._obtener_nombre_dia(col_dia)
                print(f"✅ Asignado {turno} a {elegido} en {nombre_dia}")
                return elegido

        # Si llegamos aquí, no se pudo asignar por equidad
        razones = self._analizar_razones_no_asignacion(col_dia, turno)
        self._mostrar_alerta_dia_no_asignado(col_dia, turno, razones)
        return None

    def _aplicar_formato_turno(self, fila: int, col_dia: int, turno: str) -> None:
        """Aplica formato naranja claro a la celda asignada con turno 1T o 7."""
        celda = self.ws.cell(row=fila, column=col_dia)
        celda.fill = self.fill_naranja_claro

    def _formatear_turnos_existentes(self) -> None:
        """Aplica formato naranja claro a todos los turnos 1T y 7 ya existentes en la hoja."""
        max_col = self.ws.max_column
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is not None:
                    val = str(valor).strip().upper()
                    if val in {"1T", "7"}:
                        self._aplicar_formato_turno(fila, col, val)

    def _mostrar_resumen_final(self) -> None:
        """Muestra un resumen final de todos los días no asignados."""
        if not self.dias_no_asignados:
            print(f"\n🎉 ¡EXCELENTE! Todos los días elegibles fueron asignados correctamente.")
            return
        
        print(f"\n{'='*80}")
        print(f"📊 RESUMEN FINAL - DÍAS NO ASIGNADOS")
        print(f"{'='*80}")
        print(f"Total de días no asignados: {len(self.dias_no_asignados)}")
        
        # Agrupar por razón principal
        por_razon = defaultdict(list)
        for dia in self.dias_no_asignados:
            if dia['ya_existe_turno']:
                por_razon['Ya existe turno 1T/7'].append(dia)
            elif dia['operativos'] <= 8:
                por_razon['Personal insuficiente'].append(dia)
            else:
                por_razon['Restricciones de trabajadores'].append(dia)
        
        for razon, dias in por_razon.items():
            print(f"\n🔍 {razon} ({len(dias)} días):")
            for dia in dias:
                print(f"   • {dia['dia']} (Col {dia['columna']}) - Objetivo: {dia['turno_objetivo']}")

    def procesar_todos_los_dias(self) -> None:
        print(f"🚀 Iniciando asignación de turnos 1T/7...")
        print(f"📁 Archivo: {self.archivo_procesado}")
        
        # Formatear turnos 1T/7 ya existentes antes de procesar
        print(f"🎨 Aplicando formato naranja claro a turnos 1T/7 existentes...")
        self._formatear_turnos_existentes()
        
        max_col = self.ws.max_column
        asignaciones_exitosas = 0
        
        for col in range(2, max_col + 1):
            resultado = self.asignar_turno_en_dia(col)
            if resultado:
                asignaciones_exitosas += 1

        self._actualizar_hoja_estadisticas()
        self._mostrar_resumen_final()

        salida = "horarioUnificado_con_1t.xlsx"
        self.wb.save(salida)
        print(f"\n💾 Archivo guardado como: {salida}")
        print(f"📈 Asignaciones exitosas: {asignaciones_exitosas}/{max_col-1} días")
        print(f"🎨 Celdas con turnos 1T/7 formateadas con color naranja claro")

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")   # 1T + 7
        ws_stats.cell(row=1, column=4, value="6RT")  # Solo 7
        ws_stats.cell(row=1, column=5, value="1D")   # BANTD + BLPTD
        ws_stats.cell(row=1, column=6, value="3D")   # 3 + 3D
        ws_stats.cell(row=1, column=7, value="6D")   # NLPTD + NLPRD + NANTD + NANRD

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, 8):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        # Filas de trabajadores desde la hoja principal
        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP (fórmula dinámica)
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7 (fórmula dinámica)
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = solo 7 (fórmula dinámica)
            formula_6rt = f'=COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 1D = BANTD + BLPTD
            formula_1d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"BANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"BLPTD")'
            )
            ws_stats.cell(row=fila_destino, column=5, value=formula_1d)

            # 3D = 3 + 3D
            formula_3d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"3")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"3D")'
            )
            ws_stats.cell(row=fila_destino, column=6, value=formula_3d)

            # 6D = NLPTD + NLPRD + NANTD + NANRD
            formula_6d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"NLPTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NLPRD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANRD")'
            )
            ws_stats.cell(row=fila_destino, column=7, value=formula_6d)

            fila_destino += 1

        # Anchos de columna
        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 8
        ws_stats.column_dimensions['C'].width = 8
        ws_stats.column_dimensions['D'].width = 8
        ws_stats.column_dimensions['E'].width = 8
        ws_stats.column_dimensions['F'].width = 8
        ws_stats.column_dimensions['G'].width = 8


if __name__ == "__main__":
    asignador = AsignadorTurnos()
    asignador.procesar_todos_los_dias() 