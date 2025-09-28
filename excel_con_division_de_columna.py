import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import copy

def generar_reporte_turnos(ws):
    """
    Genera un reporte de todos los turnos encontrados en el archivo
    y su estado de conversión.
    """
    turnos_encontrados = {}
    turnos_convertidos = {
        # Turnos que ocupan dos columnas (original → primera/segunda)
        "6TT": "TLPT/NLPT",
        "6RT": "MLPR/NLPR",
        "6T": "TANT/NANT",
        "6R": "MAST/NANR",
        "6N": "MANR/TANR",
        "6S": "MASR/TASR",
        "6MT": "MLPR/TLPR",
        "3": "TAST/SLN3",
        "3D": "TAST/SLN3",
        "7": "BLPT/NLPR",
        "MANRAS": "MANR/ASIG",
        "MCORTS": "MCOR/TASA",
        "N": "MANA/TANA",
        "S": "MASA/TASA",
        
        # Turnos que ocupan solo la primera columna (se mantienen o se convierten)
        "1T": "BLPT",
        "1": "BANT",
        "BANTD": "BANT",
        "BLPTD": "BLPT",
        "CMED": "CMED",
        "COME": "COME",
        "COMS": "COMS",
        "DESC": "DESC",
        "LIBR": "LIBR",
        "MANR": "MANR",
        "MASR": "MASR",
        "MN": "MANA",
        "MS": "MASA",
        "NANTD": "NANT",
        "NANRD": "NANR",
        "NLPRD": "NLPR",
        "NLPTD": "NLPT",
        "TN": "TANA",
        "TS": "TASA",
    }
    
    # Recolectar todos los turnos únicos
    for fila in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            valor = ws.cell(row=fila, column=col).value
            if valor:
                turno = str(valor).strip()
                if turno not in turnos_encontrados:
                    turnos_encontrados[turno] = 0
                turnos_encontrados[turno] += 1
    
    # Generar el reporte
    with open('reporte_conversion_turnos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE CONVERSIÓN DE TURNOS\n")
        f.write("==============================\n\n")
        
        f.write("1. TURNOS CON CONVERSIÓN DEFINIDA:\n")
        f.write("--------------------------------\n")
        f.write("a) Turnos que ocupan dos columnas:\n")
        for turno, conversion in turnos_convertidos.items():
            if '/' in conversion:
                cantidad = turnos_encontrados.get(turno, 0)
                f.write(f"* {turno:<4} → {conversion:<10} (Encontrado {cantidad} veces)\n")
        
        f.write("\nb) Turnos que ocupan una columna:\n")
        for turno, conversion in turnos_convertidos.items():
            if '/' not in conversion:
                cantidad = turnos_encontrados.get(turno, 0)
                f.write(f"* {turno:<4} → {conversion:<10} (Encontrado {cantidad} veces)\n")
        
        f.write("\n2. TURNOS SIN CONVERSIÓN DEFINIDA:\n")
        f.write("--------------------------------\n")
        for turno, cantidad in turnos_encontrados.items():
            if turno not in turnos_convertidos:
                f.write(f"* {turno:<4} (Encontrado {cantidad} veces)\n")
        
        # Estadísticas generales
        total_turnos = sum(turnos_encontrados.values())
        turnos_con_conversion = sum(turnos_encontrados.get(turno, 0) for turno in turnos_convertidos)
        
        f.write("\n3. ESTADÍSTICAS GENERALES:\n")
        f.write("-------------------------\n")
        f.write(f"Total de turnos en el archivo: {total_turnos}\n")
        f.write(f"Turnos con conversión definida: {turnos_con_conversion}\n")
        f.write(f"Turnos sin conversión definida: {total_turnos - turnos_con_conversion}\n")
        f.write(f"Porcentaje de cobertura: {(turnos_con_conversion/total_turnos*100):.2f}%\n")
    
    print("\nReporte de conversión de turnos generado en: reporte_conversion_turnos.txt")

def verificar_turnos_repetidos(ws):
    """
    Verifica que no haya turnos repetidos en el mismo día.
    Genera un reporte detallado de cualquier repetición encontrada.
    Ignora ciertos turnos que pueden repetirse (DESC, TROP, VACA, LIBR, SIND).
    """
    # Lista de turnos que pueden repetirse
    turnos_permitidos_repetir = {"DESC", "TROP", "VACA", "LIBR", "SIND"}
    
    with open('reporte_turnos_repetidos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE VERIFICACIÓN DE TURNOS REPETIDOS POR DÍA\n")
        f.write("===============================================\n\n")
        f.write("Nota: Los siguientes turnos están excluidos del análisis de repeticiones:\n")
        for turno in sorted(turnos_permitidos_repetir):
            f.write(f"- {turno}\n")
        f.write("\n")
        
        hay_repetidos = False
        # Para cada columna (día) excepto la primera que contiene las siglas
        for col in range(2, ws.max_column + 1):
            dia = ws.cell(row=1, column=col).value
            turnos_del_dia = {}
            
            # Revisar cada turno en la columna
            for fila in range(2, ws.max_row + 1):
                turno = ws.cell(row=fila, column=col).value
                if turno:  # Solo procesar celdas no vacías
                    turno = str(turno).strip()
                    # Solo registrar turnos que no están en la lista de permitidos
                    if turno not in turnos_permitidos_repetir:
                        if turno not in turnos_del_dia:
                            turnos_del_dia[turno] = []
                        turnos_del_dia[turno].append(fila)
            
            # Verificar si hay repeticiones
            turnos_repetidos = {turno: filas for turno, filas in turnos_del_dia.items() if len(filas) > 1}
            
            if turnos_repetidos:
                hay_repetidos = True
                f.write(f"\nDía: {dia}\n")
                f.write("-" * (len(str(dia)) + 5) + "\n")
                for turno, filas in turnos_repetidos.items():
                    trabajadores = [ws.cell(row=fila, column=1).value for fila in filas]
                    f.write(f"* Turno {turno} repetido {len(filas)} veces:\n")
                    for i, (fila, trabajador) in enumerate(zip(filas, trabajadores), 1):
                        f.write(f"  {i}. Fila {fila} - Trabajador: {trabajador}\n")
                f.write("\n")
        
        if not hay_repetidos:
            f.write("\nNo se encontraron turnos repetidos en ningún día.\n")
            f.write("Todos los turnos están asignados correctamente sin duplicados.\n")
            f.write("(Excluyendo los turnos permitidos para repetirse)\n")
        
        print("\nReporte de verificación de turnos repetidos generado en: reporte_turnos_repetidos.txt")
        return hay_repetidos

def verificar_cobertura_turnos(ws):
    """
    Verifica que todos los días tengan los turnos necesarios.
    Genera un reporte detallado de la cobertura de turnos.
    """
    # Definir los turnos que deben estar presentes cada día
    turnos_requeridos = {
        # Turnos que se dividen en dos columnas
        "TLPT/NLPT": 0,  # convertido de 6TT
        "MLPR/NLPR": 0,  # convertido de 6RT
        "TANT/NANT": 0,  # convertido de 6T
        "MAST/NANR": 0,  # convertido de 6R
        "MANR/TANR": 0,  # convertido de 6N
        "MASR/TASR": 0,  # convertido de 6S
        "TAST/SLN3": 0,  # convertido de 3
        
        # Turnos que ocupan una columna
        "BLPT": 0,      # convertido de 1T o BLPTD
        "BANT": 0,      # convertido de 1 o BANTD
        "MANA": 0,      # convertido de MN
        "MASA": 0,      # convertido de MS
        "TANA": 0,      # convertido de TN
        "TASA": 0       # convertido de TS
    }
    
    with open('reporte_cobertura_turnos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE COBERTURA DE TURNOS POR DÍA\n")
        f.write("=====================================\n\n")
        
        # Para cada columna (día) excepto la primera que contiene las siglas
        for col in range(2, ws.max_column + 1, 2):  # Avanzamos de 2 en 2 por las columnas divididas
            dia = ws.cell(row=1, column=col).value
            if not dia:  # Si no hay día, saltamos
                continue
                
            # Reiniciar contadores para este día
            turnos_encontrados = turnos_requeridos.copy()
            turnos_adicionales = {}
            
            # Revisar cada turno en la columna
            for fila in range(2, ws.max_row + 1):
                turno_primera = ws.cell(row=fila, column=col).value
                turno_segunda = ws.cell(row=fila, column=col + 1).value
                
                if turno_primera and turno_segunda:
                    # Es un turno dividido
                    turno_completo = f"{turno_primera}/{turno_segunda}"
                    if turno_completo in turnos_encontrados:
                        turnos_encontrados[turno_completo] += 1
                    else:
                        if turno_completo not in turnos_adicionales:
                            turnos_adicionales[turno_completo] = 0
                        turnos_adicionales[turno_completo] += 1
                elif turno_primera:
                    # Es un turno de una columna
                    if turno_primera in turnos_encontrados:
                        turnos_encontrados[turno_primera] += 1
                    else:
                        if turno_primera not in turnos_adicionales:
                            turnos_adicionales[turno_primera] = 0
                        turnos_adicionales[turno_primera] += 1
            
            # Escribir resultados para este día
            f.write(f"\nDía: {dia}\n")
            f.write("-" * (len(str(dia)) + 5) + "\n")
            
            # Turnos faltantes
            turnos_faltantes = [turno for turno, count in turnos_encontrados.items() if count == 0]
            if turnos_faltantes:
                f.write("\nTurnos faltantes:\n")
                for turno in sorted(turnos_faltantes):
                    f.write(f"- {turno}\n")
            
            # Turnos presentes
            f.write("\nTurnos presentes:\n")
            for turno, count in sorted(turnos_encontrados.items()):
                if count > 0:
                    f.write(f"- {turno}: {count} vez(ces)\n")
            
            # Turnos adicionales (no requeridos pero presentes)
            if turnos_adicionales:
                f.write("\nTurnos adicionales encontrados:\n")
                for turno, count in sorted(turnos_adicionales.items()):
                    f.write(f"- {turno}: {count} vez(ces)\n")
            
            f.write("\n" + "="*40 + "\n")
        
        print("\nReporte de cobertura de turnos generado en: reporte_cobertura_turnos.txt")

def generar_resumen_turnos_por_dia(ws):
    """
    Genera un reporte resumido que muestra solo los turnos presentes cada día.
    """
    with open('resumen_turnos_por_dia.txt', 'w', encoding='utf-8') as f:
        f.write("RESUMEN DE TURNOS PRESENTES POR DÍA\n")
        f.write("===================================\n\n")
        
        # Para cada columna (día) excepto la primera que contiene las siglas
        for col in range(2, ws.max_column + 1, 2):  # Avanzamos de 2 en 2 por las columnas divididas
            dia = ws.cell(row=1, column=col).value
            if not dia:  # Si no hay día, saltamos
                continue
            
            # Conjunto para almacenar los turnos únicos de este día
            turnos_del_dia = set()
            
            # Revisar cada turno en la columna
            for fila in range(2, ws.max_row + 1):
                turno_primera = ws.cell(row=fila, column=col).value
                turno_segunda = ws.cell(row=fila, column=col + 1).value
                
                # Filtrar fórmulas y valores no válidos
                if turno_primera and isinstance(turno_primera, str):
                    if not turno_primera.startswith('='):  # No es una fórmula
                        if turno_segunda and isinstance(turno_segunda, str) and not turno_segunda.startswith('='):
                            # Es un turno dividido
                            turno_completo = f"{turno_primera}/{turno_segunda}"
                            turnos_del_dia.add(turno_completo)
                        else:
                            # Es un turno de una columna
                            turnos_del_dia.add(turno_primera)
            
            # Escribir resultados para este día
            f.write(f"\nDía: {dia}\n")
            f.write("-" * (len(str(dia)) + 5) + "\n")
            
            # Escribir turnos ordenados alfabéticamente
            for turno in sorted(turnos_del_dia):
                f.write(f"- {turno}\n")
            
            f.write("\n" + "="*40 + "\n")
        
        print("\nResumen de turnos por día generado en: resumen_turnos_por_dia.txt")

def verificar_turnos_requeridos(ws):
    """
    Verifica que los turnos requeridos estén presentes cada día.
    """
    # Lista de turnos que deben estar presentes cada día
    turnos_requeridos = {
        # Grupo 1
        "BLPT",
        "MLPR",
        "NLPR",
        "TLPT",
        "NLPT",
        "TLPR",
        
        # Grupo 2
        "BANT",
        "MAST",
        "NANR",
        "TANT",
        "NANT",
        "TAST/SLN3",
        "MANR",
        "MASR",
        "TANR",
        "TASR"
    }
    
    # Turnos que cuentan como otros turnos
    turnos_equivalentes = {
        "MAST/NANR": ["MAST", "NANR"],
        "TANT/NANT": ["TANT", "NANT"],
        "MASR/TASR": ["MASR", "TASR"],
        "MANR/TANR": ["MANR", "TANR"],
        "MANR/ASIG": ["MANR"],
        "MLPR/NLPR": ["MLPR", "NLPR"],
        "TLPT/NLPT": ["TLPT", "NLPT"],
        "BLPT/NLPR": ["BLPT", "NLPR"],
        "MLPR/TLPR": ["MLPR", "TLPR"]

    }
    
    with open('reporte_turnos_requeridos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE VERIFICACIÓN DE TURNOS REQUERIDOS POR DÍA\n")
        f.write("================================================\n\n")
        
        # Para cada columna (día) excepto la primera que contiene las siglas
        for col in range(2, ws.max_column + 1, 2):  # Avanzamos de 2 en 2 por las columnas divididas
            dia = ws.cell(row=1, column=col).value
            if not dia:  # Si no hay día, saltamos
                continue
                
            # Conjunto para almacenar los turnos encontrados este día
            turnos_encontrados = set()
            
            # Revisar cada turno en la columna
            for fila in range(2, ws.max_row + 1):
                turno_primera = ws.cell(row=fila, column=col).value
                turno_segunda = ws.cell(row=fila, column=col + 1).value
                
                if turno_primera and isinstance(turno_primera, str):
                    if not turno_primera.startswith('='):  # No es una fórmula
                        if turno_segunda and isinstance(turno_segunda, str) and not turno_segunda.startswith('='):
                            # Es un turno dividido
                            turno_completo = f"{turno_primera}/{turno_segunda}"
                            # Si el turno dividido cuenta como turnos individuales, los agregamos
                            if turno_completo in turnos_equivalentes:
                                for turno_equiv in turnos_equivalentes[turno_completo]:
                                    turnos_encontrados.add(turno_equiv)
                            else:
                                turnos_encontrados.add(turno_completo)
                        else:
                            # Es un turno de una columna
                            turnos_encontrados.add(turno_primera)
            
            # Verificar turnos faltantes
            turnos_faltantes = turnos_requeridos - turnos_encontrados
            
            # Escribir resultados para este día
            f.write(f"\nDía: {dia}\n")
            f.write("-" * (len(str(dia)) + 5) + "\n")
            
            if turnos_faltantes:
                f.write("❌ TURNOS FALTANTES:\n")
                for turno in sorted(turnos_faltantes):
                    f.write(f"- {turno}\n")
                
                f.write("\nTurnos encontrados:\n")
                for turno in sorted(turnos_encontrados):
                    if turno in turnos_requeridos:
                        f.write(f"✓ {turno}\n")
                    else:
                        f.write(f"  {turno}\n")
            else:
                f.write("✅ Todos los turnos requeridos están presentes.\n")
            
            f.write("\n" + "="*40 + "\n")
        
        print("\nReporte de verificación de turnos requeridos generado en: reporte_turnos_requeridos.txt")

def verificar_turnos_consecutivos_prohibidos(ws):
    """
    Verifica que trabajadores con turno NLPR, NANR o NLPT no tengan BANT o BLPT al día siguiente.
    Genera un reporte detallado de cualquier violación encontrada.
    """
    with open('reporte_turnos_consecutivos_prohibidos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE VERIFICACIÓN DE TURNOS CONSECUTIVOS PROHIBIDOS\n")
        f.write("======================================================\n\n")
        f.write("Verificando que trabajadores con turno NLPR, NANR o NLPT no tengan BANT o BLPT al día siguiente.\n")
        f.write("NOTA: Para NLPT, la verificación solo se aplica si el día actual incluye el turno combinado 'BLPT/NLPR'.\n\n")
        
        violaciones_encontradas = []
        
        # Para cada trabajador (fila)
        for fila in range(2, ws.max_row + 1):
            trabajador = ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
                
            # Para cada día (columna), avanzando de 2 en 2 por las columnas divididas
            for col in range(2, ws.max_column - 1, 2):  # -1 para no ir al último día
                dia_actual = ws.cell(row=1, column=col).value
                dia_siguiente = ws.cell(row=1, column=col + 2).value
                
                if not dia_actual or not dia_siguiente:
                    continue
                
                # Obtener turnos del día actual
                turno_primera_actual = ws.cell(row=fila, column=col).value
                turno_segunda_actual = ws.cell(row=fila, column=col + 1).value
                
                # Obtener turnos del día siguiente
                turno_primera_siguiente = ws.cell(row=fila, column=col + 2).value
                turno_segunda_siguiente = ws.cell(row=fila, column=col + 3).value
                
                # Verificar si el trabajador tiene NLPR, NANR o NLPT en el día actual
                tiene_turno_prohibido = False
                turno_actual = ""
                
                if turno_primera_actual and isinstance(turno_primera_actual, str):
                    if not turno_primera_actual.startswith('='):  # No es una fórmula
                        if turno_segunda_actual and isinstance(turno_segunda_actual, str) and not turno_segunda_actual.startswith('='):
                            # Es un turno dividido
                            turno_completo = f"{turno_primera_actual}/{turno_segunda_actual}"
                            if "NLPR" in turno_completo or "NANR" in turno_completo or "NLPT" in turno_completo:
                                tiene_turno_prohibido = True
                                turno_actual = turno_completo
                        else:
                            # Es un turno de una columna
                            if turno_primera_actual in ["NLPR", "NANR", "NLPT"]:
                                tiene_turno_prohibido = True
                                turno_actual = turno_primera_actual
                
                # Si tiene NLPR, NANR o NLPT, verificar el día siguiente
                if tiene_turno_prohibido:
                    # Verificar si el día actual incluye el turno combinado "BLPT/NLPR"
                    dia_actual_tiene_blpt_nlpr = False
                    
                    # Buscar en todas las filas del día actual si existe BLPT/NLPR
                    for fila_busqueda in range(2, ws.max_row + 1):
                        turno_primera_busqueda = ws.cell(row=fila_busqueda, column=col).value
                        turno_segunda_busqueda = ws.cell(row=fila_busqueda, column=col + 1).value
                        
                        if (turno_primera_busqueda and isinstance(turno_primera_busqueda, str) and 
                            turno_segunda_busqueda and isinstance(turno_segunda_busqueda, str) and
                            not turno_primera_busqueda.startswith('=') and 
                            not turno_segunda_busqueda.startswith('=')):
                            
                            turno_completo_busqueda = f"{turno_primera_busqueda}/{turno_segunda_busqueda}"
                            if turno_completo_busqueda == "BLPT/NLPR":
                                dia_actual_tiene_blpt_nlpr = True
                                break
                    
                    # Solo verificar BANT/BLPT si:
                    # 1. El turno actual es NLPR o NANR (siempre se verifica)
                    # 2. El turno actual es NLPT Y el día actual tiene BLPT/NLPR
                    debe_verificar = False
                    if "NLPR" in turno_actual or "NANR" in turno_actual:
                        debe_verificar = True
                    elif "NLPT" in turno_actual and dia_actual_tiene_blpt_nlpr:
                        debe_verificar = True
                    
                    if debe_verificar:
                        tiene_bant_o_blpt = False
                        turno_siguiente = ""
                        
                        if turno_primera_siguiente and isinstance(turno_primera_siguiente, str):
                            if not turno_primera_siguiente.startswith('='):  # No es una fórmula
                                if turno_segunda_siguiente and isinstance(turno_segunda_siguiente, str) and not turno_segunda_siguiente.startswith('='):
                                    # Es un turno dividido
                                    turno_completo = f"{turno_primera_siguiente}/{turno_segunda_siguiente}"
                                    if "BANT" in turno_completo or "BLPT" in turno_completo:
                                        tiene_bant_o_blpt = True
                                        turno_siguiente = turno_completo
                                else:
                                    # Es un turno de una columna
                                    if turno_primera_siguiente in ["BANT", "BLPT"]:
                                        tiene_bant_o_blpt = True
                                        turno_siguiente = turno_primera_siguiente
                        
                        # Si hay violación, agregarla a la lista
                        if tiene_bant_o_blpt:
                            violaciones_encontradas.append({
                                'trabajador': trabajador,
                                'dia_actual': dia_actual,
                                'dia_siguiente': dia_siguiente,
                                'turno_actual': turno_actual,
                                'turno_siguiente': turno_siguiente,
                                'fila': fila,
                                'dia_actual_tiene_blpt_nlpr': dia_actual_tiene_blpt_nlpr
                            })
        
        # Escribir el reporte
        if violaciones_encontradas:
            f.write(f"❌ SE ENCONTRARON {len(violaciones_encontradas)} VIOLACIONES:\n\n")
            
            for i, violacion in enumerate(violaciones_encontradas, 1):
                f.write(f"{i}. TRABAJADOR: {violacion['trabajador']}\n")
                f.write(f"   Día actual ({violacion['dia_actual']}): {violacion['turno_actual']}\n")
                f.write(f"   Día siguiente ({violacion['dia_siguiente']}): {violacion['turno_siguiente']}\n")
                f.write(f"   Fila en Excel: {violacion['fila']}\n")
                if 'dia_actual_tiene_blpt_nlpr' in violacion:
                    f.write(f"   Día actual tiene BLPT/NLPR: {'Sí' if violacion['dia_actual_tiene_blpt_nlpr'] else 'No'}\n")
                f.write(f"   ⚠️  VIOLACIÓN: {violacion['turno_actual']} seguido de {violacion['turno_siguiente']}\n\n")
            
            f.write("="*60 + "\n")
            f.write("RECOMENDACIONES:\n")
            f.write("- Revisar la asignación de turnos para estos trabajadores\n")
            f.write("- Considerar asignar turnos de descanso o turnos diferentes\n")
            f.write("- Verificar que se cumplan las reglas de descanso entre turnos\n")
        else:
            f.write("✅ NO SE ENCONTRARON VIOLACIONES\n\n")
            f.write("Todos los trabajadores que tuvieron turno NLPR, NANR o NLPT\n")
            f.write("NO tienen turno BANT o BLPT al día siguiente.\n")
            f.write("Las asignaciones cumplen con las reglas establecidas.\n")
        
        print(f"\nReporte de verificación de turnos consecutivos prohibidos generado en: reporte_turnos_consecutivos_prohibidos.txt")
        print(f"Violaciones encontradas: {len(violaciones_encontradas)}")
        
        return len(violaciones_encontradas) > 0

def contar_repeticiones_turnos_especificos(ws):
    """
    Cuenta las repeticiones de turnos específicos en todo el archivo.
    Genera un reporte detallado con la cantidad de veces que aparece cada turno.
    """
    # Lista de turnos específicos a contar
    turnos_especificos = {
        "BLPT": 0,
        "MLPR": 0,
        "NLPR": 0,
        "TLPT": 0,
        "NLPT": 0,
        "TLPR": 0,
        "BANT": 0,
        "MAST": 0,
        "NANR": 0,
        "TANT": 0,
        "NANT": 0,
        "TAST/SLN3": 0,
        "MANR": 0,
        "MASR": 0,
        "TANR": 0,
        "TASR": 0
    }
    
    # Turnos que cuentan como otros turnos (turnos divididos)
    turnos_equivalentes = {
        "MAST/NANR": ["MAST", "NANR"],
        "TANT/NANT": ["TANT", "NANT"],
        "MASR/TASR": ["MASR", "TASR"],
        "MANR/TANR": ["MANR", "TANR"],
        "MANR/ASIG": ["MANR"],
        "MLPR/NLPR": ["MLPR", "NLPR"],
        "TLPT/NLPT": ["TLPT", "NLPT"],
        "BLPT/NLPR": ["BLPT", "NLPR"],
        "MLPR/TLPR": ["MLPR", "TLPR"],
        "TAST/SLN3": ["TAST/SLN3"]
    }
    
    with open('reporte_repeticiones_turnos_especificos.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE REPETICIONES DE TURNOS ESPECÍFICOS\n")
        f.write("==========================================\n\n")
        f.write("Conteo de las veces que aparecen los siguientes turnos en todo el archivo:\n\n")
        
        # Para cada columna (día) excepto la primera que contiene las siglas
        for col in range(2, ws.max_column + 1, 2):  # Avanzamos de 2 en 2 por las columnas divididas
            dia = ws.cell(row=1, column=col).value
            if not dia:  # Si no hay día, saltamos
                continue
            
            # Revisar cada turno en la columna
            for fila in range(2, ws.max_row + 1):
                turno_primera = ws.cell(row=fila, column=col).value
                turno_segunda = ws.cell(row=fila, column=col + 1).value
                
                if turno_primera and isinstance(turno_primera, str):
                    if not turno_primera.startswith('='):  # No es una fórmula
                        if turno_segunda and isinstance(turno_segunda, str) and not turno_segunda.startswith('='):
                            # Es un turno dividido
                            turno_completo = f"{turno_primera}/{turno_segunda}"
                            
                            # Si el turno dividido cuenta como turnos individuales, los agregamos
                            if turno_completo in turnos_equivalentes:
                                for turno_equiv in turnos_equivalentes[turno_completo]:
                                    if turno_equiv in turnos_especificos:
                                        turnos_especificos[turno_equiv] += 1
                            else:
                                # Verificar si el turno completo está en la lista
                                if turno_completo in turnos_especificos:
                                    turnos_especificos[turno_completo] += 1
                        else:
                            # Es un turno de una columna
                            if turno_primera in turnos_especificos:
                                turnos_especificos[turno_primera] += 1
        
        # Escribir resultados
        f.write("RESULTADOS DEL CONTEO:\n")
        f.write("=====================\n\n")
        
        # Agrupar por categorías
        f.write("GRUPO 1 - Turnos de Línea Principal:\n")
        f.write("-----------------------------------\n")
        grupo1 = ["BLPT", "MLPR", "NLPR", "TLPT", "NLPT", "TLPR"]
        for turno in grupo1:
            cantidad = turnos_especificos[turno]
            f.write(f"* {turno:<8}: {cantidad:>3} veces\n")
        
        f.write("\nGRUPO 2 - Turnos de Análisis y Soporte:\n")
        f.write("--------------------------------------\n")
        grupo2 = ["BANT", "MAST", "NANR", "TANT", "NANT", "TAST/SLN3", "MANR", "MASR", "TANR", "TASR"]
        for turno in grupo2:
            cantidad = turnos_especificos[turno]
            f.write(f"* {turno:<10}: {cantidad:>3} veces\n")
        
        # Estadísticas generales
        total_repeticiones = sum(turnos_especificos.values())
        turnos_con_repeticiones = sum(1 for cantidad in turnos_especificos.values() if cantidad > 0)
        turnos_sin_repeticiones = len(turnos_especificos) - turnos_con_repeticiones
        
        f.write("\n" + "="*50 + "\n")
        f.write("ESTADÍSTICAS GENERALES:\n")
        f.write("======================\n")
        f.write(f"Total de repeticiones encontradas: {total_repeticiones}\n")
        f.write(f"Turnos con repeticiones: {turnos_con_repeticiones}\n")
        f.write(f"Turnos sin repeticiones: {turnos_sin_repeticiones}\n")
        f.write(f"Total de turnos monitoreados: {len(turnos_especificos)}\n")
        
        # Turnos más frecuentes
        turnos_ordenados = sorted(turnos_especificos.items(), key=lambda x: x[1], reverse=True)
        f.write("\nTURNOS MÁS FRECUENTES:\n")
        f.write("=====================\n")
        for i, (turno, cantidad) in enumerate(turnos_ordenados[:5], 1):
            if cantidad > 0:
                f.write(f"{i}. {turno}: {cantidad} veces\n")
        
        # Turnos sin repeticiones
        turnos_sin_apariciones = [turno for turno, cantidad in turnos_especificos.items() if cantidad == 0]
        if turnos_sin_apariciones:
            f.write("\nTURNOS SIN APARICIONES:\n")
            f.write("======================\n")
            for turno in sorted(turnos_sin_apariciones):
                f.write(f"* {turno}\n")
        
        print(f"\nReporte de repeticiones de turnos específicos generado en: reporte_repeticiones_turnos_especificos.txt")
        print(f"Total de repeticiones encontradas: {total_repeticiones}")

def modificar_horario_con_division_columna():
    """
    Modifica el archivo horarioUnificado_a_dividir.xlsx:
    1. Divide cada columna de día en dos columnas
    2. El encabezado del día cubre ambas columnas
    3. Aplica reglas específicas de renombrado:
       - Turnos que se dividen en dos partes (ocupan ambas columnas):
         * 6TT → TLPT/NLPT
         * 6RT → MLPR/NLPR
         * 6T → TANT/NANT
         * 6R → MAST/NANR
         * 6N → MANR/TANR
         * 6S → MASR/TASR
         * 6MT → MLPR/TLPR
         * 3 → TAST/SLN3
         * 3D → TAST/SLN3
         * 7 → BLPT/NLPR
         * MANRAS → MANR/ASIG
         * MCORTS → MCOR/TASA
         * N → MANA/TANA
         * S → MASA/TASA
       - Turnos que se renombran pero ocupan solo la primera columna:
         * 1T → BLPT
         * 1 → BANT
         * BANTD → BANT
         * BLPTD → BLPT
         * CMED → CMED
         * COME → COME
         * COMS → COMS
         * DESC → DESC
         * LIBR → LIBR
         * MANR → MANR
         * MASR → MASR
         * MN → MANA
         * MS → MASA
         * NANTD → NANT
         * NANRD → NANR
         * NLPRD → NLPR
         * NLPTD → NLPT
         * TN → TANA
         * TS → TASA
       - Otros turnos se mantienen iguales pero ocupan solo la primera columna
    """
    
    # Cargar el archivo original
    wb = openpyxl.load_workbook('horarioUnificado_a_dividir.xlsx')
    ws = wb.active
    
    print(f"Procesando archivo: {wb.active.title}")
    print(f"Dimensiones originales: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Verificar turnos repetidos antes de la conversión
    print("\nVerificando turnos repetidos en el archivo original...")
    hay_repetidos = verificar_turnos_repetidos(ws)
    if hay_repetidos:
        print("⚠️  ADVERTENCIA: Se encontraron turnos repetidos. Revise el reporte para más detalles.")
    
    # Generar reporte de turnos antes de la conversión
    generar_reporte_turnos(ws)
    
    # Crear nuevo workbook
    nuevo_wb = openpyxl.Workbook()
    nuevo_ws = nuevo_wb.active
    nuevo_ws.title = ws.title
    
    # Definir colores
    azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    rojo_claro = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    
    def aplicar_color_seguro(celda, color):
        """Aplica color solo si existe"""
        if color:
            celda.fill = color
    
    # Copiar la primera columna (trabajadores) sin cambios y aplicar color azul claro
    for fila in range(1, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        celda_nueva = nuevo_ws.cell(row=fila, column=1, value=valor)
        celda_nueva.fill = azul_claro
    
    # Procesar columnas de días (desde la columna 2 en adelante)
    nueva_col = 2
    for col_original in range(2, ws.max_column + 1):
        dia = ws.cell(row=1, column=col_original).value
        
        # Crear dos columnas para cada día
        col_primera = nueva_col
        col_segunda = nueva_col + 1
        
        # Combinar las dos columnas para el encabezado del día
        nuevo_ws.merge_cells(start_row=1, start_column=col_primera, 
                           end_row=1, end_column=col_segunda)
        celda_encabezado = nuevo_ws.cell(row=1, column=col_primera, value=dia)
        
        # Centrar el encabezado y aplicar color
        celda_encabezado.alignment = Alignment(horizontal='center')
        
        # Aplicar color según el tipo de día
        if dia and str(dia).startswith("SUN"):
            # Domingo: rojo claro
            celda_encabezado.fill = rojo_claro
        else:
            # Otros días: azul claro
            celda_encabezado.fill = azul_claro
        
        # Procesar cada fila para este día
        for fila in range(2, ws.max_row + 1):
            turno_original = ws.cell(row=fila, column=col_original).value
            
            if turno_original:
                turno_str = str(turno_original).strip()
                
                # Obtener el color original del turno
                celda_original = ws.cell(row=fila, column=col_original)
                
                # Aplicar reglas específicas de renombrado
                if turno_str == "6TT":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPT")
                elif turno_str == "6RT":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MLPR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPR")
                elif turno_str == "6T":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TANT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NANT")
                elif turno_str == "6R":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MAST")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NANR")
                elif turno_str == "6N":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MANR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TANR")
                elif turno_str == "6S":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MASR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TASR")
                elif turno_str == "6MT":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MLPR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TLPR")
                elif turno_str in ["3", "3D"]:
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TAST")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="SLN3")
                elif turno_str == "7":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPR")
                elif turno_str == "MANRAS":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MANR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="ASIG")
                elif turno_str == "MCORTS":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MCOR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TASA")
                elif turno_str == "N":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TANA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TASA")
                elif turno_str == "S":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MASA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TASA")
                # Turnos que se renombran pero ocupan solo la primera columna
                elif turno_str in ["1T", "BLPTD"]:
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str in ["1", "BANTD"]:
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BANT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "MN":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MANA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "MS":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MASA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "NANTD":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="NANT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "NANRD":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="NANR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "NLPRD":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="NLPR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "NLPTD":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="NLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "TN":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TANA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                elif turno_str == "TS":
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TASA")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
                else:
                    # Otros turnos se mantienen iguales pero ocupan solo la primera columna
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value=turno_original)
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")
        
        # Aplicar color rojo claro a las dos columnas de domingos
        if dia and str(dia).startswith("SUN"):
            for fila in range(2, ws.max_row + 1):
                # Aplicar color rojo claro a las celdas vacías de domingo
                celda1 = nuevo_ws.cell(row=fila, column=col_primera)
                celda2 = nuevo_ws.cell(row=fila, column=col_segunda)
                
                # Solo aplicar color si la celda no tiene contenido (está vacía)
                if not celda1.value:
                    celda1.fill = rojo_claro
                if not celda2.value:
                    celda2.fill = rojo_claro
        
        nueva_col += 2  # Avanzar dos columnas para el siguiente día
    
    # Ajustar ancho de columnas
    for col in range(1, nuevo_ws.max_column + 1):
        nuevo_ws.column_dimensions[get_column_letter(col)].width = 6  # Reducido de 8 a 6
    
    # Verificar cobertura de turnos antes de guardar
    print("\nVerificando cobertura de turnos en el archivo modificado...")
    verificar_cobertura_turnos(nuevo_ws)
    
    # Generar resumen de turnos por día
    print("\nGenerando resumen de turnos por día...")
    generar_resumen_turnos_por_dia(nuevo_ws)
    
    # Verificar turnos requeridos
    print("\nVerificando presencia de turnos requeridos...")
    verificar_turnos_requeridos(nuevo_ws)
    
    # Verificar turnos consecutivos prohibidos
    print("\nVerificando turnos consecutivos prohibidos...")
    hay_violaciones = verificar_turnos_consecutivos_prohibidos(nuevo_ws)
    if hay_violaciones:
        print("⚠️  ADVERTENCIA: Se encontraron violaciones de turnos consecutivos prohibidos. Revise el reporte para más detalles.")
    
    # Contar repeticiones de turnos específicos
    print("\nContando repeticiones de turnos específicos...")
    contar_repeticiones_turnos_especificos(nuevo_ws)
    
    # Guardar el archivo modificado
    nombre_archivo_salida = "excel_con_division_de_columna.xlsx"
    try:
        nuevo_wb.save(nombre_archivo_salida)
        print(f"\nArchivo modificado guardado como: {nombre_archivo_salida}")
        print(f"Nuevas dimensiones: {nuevo_ws.max_row} filas x {nuevo_ws.max_column} columnas")
    except PermissionError:
        print(f"\n⚠️  No se pudo guardar el archivo {nombre_archivo_salida}.")
        print("   Por favor, cierre el archivo si está abierto en Excel.")
    
    # Mostrar resumen de los cambios
    print("\nResumen de cambios realizados:")
    print("- Cada columna de día se dividió en dos columnas")
    print("- El encabezado del día cubre ambas columnas")
    print("- Turnos que se dividen en dos partes (ocupan ambas columnas):")
    print("  * 6TT → TLPT/NLPT")
    print("  * 6RT → MLPR/NLPR")
    print("  * 6T → TANT/NANT")
    print("  * 6R → MAST/NANR")
    print("  * 6N → MANR/TANR")
    print("  * 6S → MASR/TASR")
    print("  * 6MT → MLPR/TLPR")
    print("  * 3 → TAST/SLN3")
    print("  * 3D → TAST/SLN3")
    print("  * 7 → BLPT/NLPR")
    print("  * MANRAS → MANR/ASIG")
    print("  * MCORTS → MCOR/TASA")
    print("  * N → MANA/TANA")
    print("  * S → MASA/TASA")
    print("- Turnos que se renombran pero ocupan solo la primera columna:")
    print("  * 1T, BLPTD → BLPT")
    print("  * 1, BANTD → BANT")
    print("  * MN → MANA")
    print("  * MS → MASA")
    print("  * NANTD → NANT")
    print("  * NANRD → NANR")
    print("  * NLPRD → NLPR")
    print("  * NLPTD → NLPT")
    print("  * TN → TANA")
    print("  * TS → TASA")
    print("- Otros turnos se mantienen iguales pero ocupan solo la primera columna")

def mostrar_estructura_archivo():
    """
    Muestra la estructura del archivo original para referencia
    """
    try:
        wb = openpyxl.load_workbook('horarioUnificado_a_dividir.xlsx')
        ws = wb.active
        
        print(f"Estructura del archivo original:")
        print(f"Archivo: horarioUnificado_a_dividir.xlsx")
        print(f"Hoja: {ws.title}")
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
        
        # Mostrar encabezados
        print("\nEncabezados (primera fila):")
        for col in range(1, min(ws.max_column + 1, 10)):  # Primeras 10 columnas
            valor = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {repr(valor)}")
        
        # Mostrar algunos trabajadores
        print("\nPrimeros trabajadores:")
        for fila in range(2, min(ws.max_row + 1, 8)):  # Primeros 6 trabajadores
            trabajador = ws.cell(row=fila, column=1).value
            print(f"  Fila {fila}: {trabajador}")
        
        # Buscar turnos específicos
        print("\nBuscando turnos específicos:")
        turnos_encontrados = set()
        for fila in range(2, ws.max_row + 1):
            for col in range(2, ws.max_column + 1):
                valor = ws.cell(row=fila, column=col).value
                if valor and str(valor).strip() in ["6TT", "6T", "6RT", "6R", "1T", "1", "7"]:
                    turnos_encontrados.add(str(valor).strip())
        
        print(f"Turnos encontrados: {sorted(turnos_encontrados)}")
        
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

if __name__ == "__main__":
    # Mostrar estructura del archivo original
    mostrar_estructura_archivo()
    
    print("\n" + "="*50)
    print("INICIANDO MODIFICACIÓN DEL ARCHIVO")
    print("="*50)
    
    # Ejecutar la modificación
    modificar_horario_con_division_columna() 